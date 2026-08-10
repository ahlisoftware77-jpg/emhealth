import os
import re
import io
import time
import smtplib
from typing import List, Optional
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, BackgroundTasks
from pydantic import BaseModel
from PIL import Image, ImageOps
import openpyxl

# Register HEIC opener if available
try:
    import pillow_heif
    pillow_heif.register_heif_opener()
    HAS_HEIF = True
except ImportError:
    HAS_HEIF = False

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate, make_msgid

router = APIRouter(prefix="/mcu-blast", tags=["MCU Blast & Image Optimizer"])

STORAGE_DIR = r"D:\COMPRESS"
INPUT_DIR = os.path.join(STORAGE_DIR, "foto asli")
OUTPUT_DIR = os.path.join(STORAGE_DIR, "output")
DEFAULT_EXCEL_PATH = os.path.join(STORAGE_DIR, "idpass-update2.xlsx")

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

class SMTPVerifyRequest(BaseModel):
    mode: str = "GMAIL" # "GMAIL" or "CUSTOM_DOMAIN"
    sender_email: str
    sender_password: str
    smtp_server: Optional[str] = "smtp.gmail.com"
    smtp_port: Optional[int] = 587
    use_ssl: Optional[bool] = False

class RecipientItem(BaseModel):
    Nama: str
    UseriD: str
    Password: str
    Email: str

class SendEmailsRequest(BaseModel):
    mode: str = "GMAIL"
    sender_email: str
    sender_password: str
    smtp_server: Optional[str] = "smtp.gmail.com"
    smtp_port: Optional[int] = 587
    use_ssl: Optional[bool] = False
    sender_name: Optional[str] = "EM-Health Admin"
    program_name: Optional[str] = "Medical Check Up"
    company_name: Optional[str] = "PT. YADIKOMPUTER"
    link_url: Optional[str] = "https://mcu-emhealth.com/login-employee"
    delay_seconds: Optional[int] = 5
    dry_run: Optional[bool] = False
    recipients: Optional[List[RecipientItem]] = None

@router.post("/verify-smtp")
async def verify_smtp(req: SMTPVerifyRequest):
    server_host = req.smtp_server if req.mode == "CUSTOM_DOMAIN" else "smtp.gmail.com"
    server_port = req.smtp_port if req.mode == "CUSTOM_DOMAIN" else 587
    use_ssl = req.use_ssl if req.mode == "CUSTOM_DOMAIN" else False

    try:
        if use_ssl:
            server = smtplib.SMTP_SSL(server_host, server_port, timeout=10)
        else:
            server = smtplib.SMTP(server_host, server_port, timeout=10)
            server.starttls()
        
        server.login(req.sender_email, req.sender_password)
        server.quit()
        return {"success": True, "message": f"Login SMTP ({req.sender_email}) Berhasil!"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal koneksi SMTP: {str(e)}")

def parse_excel_recipients(file_bytes_or_path):
    if isinstance(file_bytes_or_path, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes_or_path), data_only=True)
    else:
        if not os.path.exists(file_bytes_or_path):
            return []
        wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True)

    sheet = wb.active
    recipients = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 2:
            nama = str(row[0]).strip() if row[0] else ""
            userid = str(row[1]).strip('"\' ') if row[1] else ""
            password = str(row[2]).strip() if len(row) >= 3 and row[2] else ""
            email = str(row[3]).strip() if len(row) >= 4 and row[3] else ""

            if userid or email:
                recipients.append({
                    "Nama": nama,
                    "UseriD": userid,
                    "Password": password,
                    "Email": email
                })
    return recipients

def match_user_id(file_name, users_list):
    if not users_list:
        return None

    name_part = os.path.splitext(file_name)[0]
    if ' - ' in name_part:
        name_part = name_part.split(' - ', 1)[1]
    elif '-' in name_part and not name_part.lower().startswith('image'):
        parts = name_part.split('-')
        if len(parts) > 1 and (parts[0].strip().isdigit() or 'IMG' in parts[0].upper()):
            name_part = parts[-1]

    fname_tokens = [t.lower() for t in re.split(r'[\s\._]+', name_part) if t and not t.isdigit()]
    fname_tokens = [t for t in fname_tokens if t not in ('image', 'img', 'photo', 'foto')]

    if not fname_tokens:
        return None

    # Step 1: Match Name (Kolom A)
    best_name_match = None
    best_name_score = 0
    for u in users_list:
        excel_name = u["Nama"]
        if not excel_name:
            continue
        name_tokens = [t.lower() for t in re.split(r'[\s\._]+', excel_name) if t]
        score = 0
        matched_words = 0
        for ft in fname_tokens:
            for nt in name_tokens:
                if ft == nt:
                    score += 20
                    matched_words += 1
                    break
                elif len(ft) == 1 and nt.startswith(ft):
                    score += 10
                    matched_words += 1
                    break
                elif len(nt) == 1 and ft.startswith(nt):
                    score += 10
                    matched_words += 1
                    break
                elif len(ft) > 2 and ft in nt:
                    score += 15
                    matched_words += 1
                    break

        if score > best_name_score and matched_words > 0:
            best_name_score = score
            best_name_match = (u["UseriD"], excel_name, "Nama Excel")

    if best_name_score >= 10:
        return best_name_match

    # Step 2: Match Email (Kolom D)
    best_email_match = None
    best_email_score = 0
    for u in users_list:
        email = u["Email"]
        if not email:
            continue
        email_clean = email.lower()
        email_user = email_clean.split('@')[0] if '@' in email_clean else email_clean
        score = 0
        matched_words = 0
        for ft in fname_tokens:
            if len(ft) <= 1:
                continue
            if ft in email_user:
                score += 20
                matched_words += 1

        if score > best_email_score and matched_words > 0:
            best_email_score = score
            best_email_match = (u["UseriD"], email, "Email Excel")

    if best_email_score >= 10:
        return best_email_match

    return None

def process_single_image(img_bytes, filename, target_bytes=500*1024):
    img = Image.open(io.BytesIO(img_bytes))
    img = ImageOps.exif_transpose(img)

    # Force Portrait
    if img.width > img.height:
        img = img.rotate(270, expand=True)

    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")

    save_format = "JPEG"
    quality = 95
    out_buf = io.BytesIO()

    while quality >= 20:
        out_buf = io.BytesIO()
        img.save(out_buf, format=save_format, optimize=True, quality=quality)
        if out_buf.tell() <= target_bytes:
            break
        quality -= 5

    if out_buf.tell() > target_bytes:
        curr_img = img.copy()
        while True:
            nw = int(curr_img.width * 0.9)
            nh = int(curr_img.height * 0.9)
            if nw < 100 or nh < 100:
                break
            curr_img = curr_img.resize((nw, nh), Image.Resampling.LANCZOS)
            out_buf = io.BytesIO()
            curr_img.save(out_buf, format=save_format, optimize=True, quality=35)
            if out_buf.tell() <= target_bytes:
                break

    return out_buf.getvalue(), img.width, img.height

@router.post("/process-images")
async def process_images(
    excel_file: Optional[UploadFile] = File(None),
    images: List[UploadFile] = File(...)
):
    if excel_file:
        content = await excel_file.read()
        recipients = parse_excel_recipients(content)
    else:
        recipients = parse_excel_recipients(DEFAULT_EXCEL_PATH)

    processed_results = []
    
    for img_file in images:
        img_bytes = await img_file.read()
        filename = img_file.filename
        
        comp_bytes, final_w, final_h = process_single_image(img_bytes, filename)
        orig_kb = round(len(img_bytes) / 1024, 2)
        comp_kb = round(len(comp_bytes) / 1024, 2)

        match_res = match_user_id(filename, recipients)
        if match_res:
            matched_uid, matched_detail, match_source = match_res
            ext = os.path.splitext(filename)[1].lower()
            if ext in (".png", ".heic", ".heif"):
                ext = ".jpg"
            final_name = f"{matched_uid}{ext}"
        else:
            final_name = filename
            ext = os.path.splitext(filename)[1].lower()
            if ext in (".heic", ".heif"):
                final_name = os.path.splitext(filename)[0] + ".jpg"
            matched_uid, matched_detail, match_source = None, None, None

        # Save to output dir
        save_path = os.path.join(OUTPUT_DIR, final_name)
        with open(save_path, "wb") as f:
            f.write(comp_bytes)

        processed_results.append({
            "original_filename": filename,
            "final_filename": final_name,
            "original_size_kb": orig_kb,
            "compressed_size_kb": comp_kb,
            "dimensions": f"{final_w}x{final_h}",
            "orientation": "Portrait",
            "matched_userid": matched_uid,
            "matched_detail": matched_detail,
            "match_source": match_source,
            "saved_path": save_path
        })

    return {
        "success": True,
        "total_recipients_in_excel": len(recipients),
        "total_images_processed": len(processed_results),
        "recipients": recipients,
        "results": processed_results
    }

@router.post("/send-emails")
async def send_emails(req: SendEmailsRequest):
    recipients = req.recipients
    if not recipients:
        recipients = [RecipientItem(**r) for r in parse_excel_recipients(DEFAULT_EXCEL_PATH)]

    if not recipients:
        raise HTTPException(status_code=400, detail="Tidak ada daftar penerima email!")

    server_host = req.smtp_server if req.mode == "CUSTOM_DOMAIN" else "smtp.gmail.com"
    server_port = req.smtp_port if req.mode == "CUSTOM_DOMAIN" else 587
    use_ssl = req.use_ssl if req.mode == "CUSTOM_DOMAIN" else False

    email_results = []
    
    if req.dry_run:
        for idx, r in enumerate(recipients, 1):
            att_name = f"{r.UseriD}.jpg"
            att_path = os.path.join(OUTPUT_DIR, att_name)
            has_att = os.path.exists(att_path)
            email_results.append({
                "index": idx,
                "nama": r.Nama,
                "email": r.Email,
                "userid": r.UseriD,
                "status": "SIMULASI OK",
                "has_attachment": has_att,
                "message": "Simulasi kirim berhasil (Dry Run)."
            })
        return {
            "success": True,
            "mode": "DRY_RUN",
            "total": len(recipients),
            "results": email_results
        }

    # Real Sending
    try:
        if use_ssl:
            server = smtplib.SMTP_SSL(server_host, server_port, timeout=15)
        else:
            server = smtplib.SMTP(server_host, server_port, timeout=15)
            server.starttls()
        server.login(req.sender_email, req.sender_password)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal koneksi SMTP: {str(e)}")

    EMAIL_HTML_TEMPLATE = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; }}
            .header {{ background-color: #007bff; color: white; padding: 20px; text-align: center; border-radius: 6px 6px 0 0; }}
            .header h1 {{ margin: 0; font-size: 24px; font-weight: bold; letter-spacing: 1px; }}
            .content {{ padding: 25px; background-color: #f9f9f9; text-align: center; }}
            .section-title {{ font-size: 18px; font-weight: bold; color: #007bff; margin-top: 15px; margin-bottom: 10px; text-align: center; }}
            .card {{ background: white; padding: 20px; border-radius: 6px; border: 1px solid #007bff; margin: 20px auto; text-align: left; max-width: 480px; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
            .footer {{ text-align: center; padding: 15px; font-size: 12px; color: #777; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>EM-HEALTH</h1>
            </div>
            <div class="content">
                <br><br>
                <div class="section-title">Hasil Pemeriksaan MCU</div>
                <br>
                <p>Halo <b>{Nama}</b>,</p>
                <br>
                <p>Hasil pemeriksaan <b>{ProgramName}</b>, <b>{CompanyName}</b> anda telah tersedia.</p>
                <br>
                <p><b>Informasi Login Peserta</b></p>
                
                <div class="card">
                    <p style="margin: 6px 0;"><b>NIK (Username Login):</b> <code>{UseriD}</code></p>
                    <p style="margin: 6px 0;"><b>Password:</b> <code>{Password}</code></p>
                    <p style="margin: 6px 0;"><b>Email Terdaftar:</b> {Email}</p>
                </div>

                <p>Harap simpan informasi akun ini dengan baik dan jangan berikan password Anda kepada pihak lain.</p>
                
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{LinkUrl}" target="_blank" style="background-color: #28a745; color: #ffffff; padding: 12px 28px; text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 15px; display: inline-block; box-shadow: 0 2px 4px rgba(0,0,0,0.15);">Unduh Hasil</a>
                </div>

                <p style="font-size: 13px; color: #666;">Jika ada foto yang dilampirkan, file tersebut telah dikompresi sesuai ketentuan sistem.</p>
            </div>
            <div class="footer">
                <p>Email ini dikirim secara resmi oleh Sistem Admin EM-Health. Harap tidak membalas email otomatis ini.</p>
            </div>
        </div>
    </body>
    </html>
    """

    try:
        for idx, r in enumerate(recipients, 1):
            msg_root = MIMEMultipart("mixed")
            msg_root["From"] = f"{req.sender_name} <{req.sender_email}>"
            msg_root["To"] = r.Email
            msg_root["Reply-To"] = req.sender_email
            msg_root["Subject"] = "Hasil Pemeriksaan MCU - EM-HEALTH"
            msg_root["Date"] = formatdate(localtime=True)
            msg_root["Message-ID"] = make_msgid(domain=req.sender_email.split('@')[-1])

            msg_alt = MIMEMultipart("alternative")

            plain_text = f"""EM-HEALTH

Hasil Pemeriksaan MCU

Halo {r.Nama},

Hasil pemeriksaan {req.program_name}, {req.company_name} anda telah tersedia.

Informasi Login Peserta:
- NIK (Username Login): {r.UseriD}
- Password: {r.Password}
- Email Terdaftar: {r.Email}

Harap simpan informasi akun ini dengan baik dan jangan berikan password Anda kepada pihak lain.

Unduh hasil pemeriksaan pada tautan berikut:
{req.link_url}

---
Email ini dikirim secara resmi oleh Sistem Admin EM-Health.
"""
            msg_alt.attach(MIMEText(plain_text, "plain", "utf-8"))

            html_body = EMAIL_HTML_TEMPLATE.format(
                Nama=r.Nama,
                UseriD=r.UseriD,
                Password=r.Password,
                Email=r.Email,
                ProgramName=req.program_name,
                CompanyName=req.company_name,
                LinkUrl=req.link_url
            )
            msg_alt.attach(MIMEText(html_body, "html", "utf-8"))
            msg_root.attach(msg_alt)

            att_filename = f"{r.UseriD}.jpg"
            att_path = os.path.join(OUTPUT_DIR, att_filename)
            has_att = False
            if os.path.exists(att_path):
                try:
                    with open(att_path, "rb") as af:
                        part = MIMEBase("image", "jpeg")
                        part.set_payload(af.read())
                    encoders.encode_base64(part)
                    part.add_header("Content-Disposition", f"attachment; filename={att_filename}")
                    msg_root.attach(part)
                    has_att = True
                except Exception as e:
                    pass

            try:
                server.sendmail(req.sender_email, r.Email, msg_root.as_string())
                email_results.append({
                    "index": idx,
                    "nama": r.Nama,
                    "email": r.Email,
                    "userid": r.UseriD,
                    "status": "SUKSES TERKIRIM",
                    "has_attachment": has_att,
                    "message": "Email berhasil dikirim."
                })
            except Exception as e:
                email_results.append({
                    "index": idx,
                    "nama": r.Nama,
                    "email": r.Email,
                    "userid": r.UseriD,
                    "status": "GAGAL",
                    "has_attachment": has_att,
                    "message": f"Gagal kirim: {str(e)}"
                })

            if idx < len(recipients) and req.delay_seconds > 0:
                time.sleep(req.delay_seconds)

    finally:
        server.quit()

    return {
        "success": True,
        "mode": "REAL",
        "total": len(recipients),
        "results": email_results
    }
