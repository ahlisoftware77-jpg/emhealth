import pandas as pd
import openpyxl
from rapidfuzz import process, fuzz
from pathlib import Path
from io import BytesIO
from typing import List, Dict, Any, Tuple, Optional
from app.core.config import settings
import logging

logger = logging.getLogger(__name__)

class ExcelService:
    @staticmethod
    def load_dataframe(file_path: Path) -> pd.DataFrame:
        ext = file_path.suffix.lower()
        if ext == ".csv":
            return pd.read_csv(file_path, dtype=str).fillna("")
        elif ext in [".xlsx", ".xls"]:
            return pd.read_excel(file_path, dtype=str).fillna("")
        else:
            raise ValueError(f"Format file tidak didukung: {ext}")

    @staticmethod
    def get_columns_and_preview_from_bytes(content: bytes, filename: str, max_rows: int = 500) -> Dict[str, Any]:
        """Membaca preview Excel/CSV langsung dari bytes di memory — tanpa menulis ke disk.
        Digunakan untuk endpoint /inspect-file agar kompatibel dengan Vercel read-only filesystem."""
        ext = Path(filename).suffix.lower()
        try:
            records = []
            columns = []
            bio = BytesIO(content)

            if ext == ".csv":
                bio.seek(0)
                try:
                    df = pd.read_csv(bio, dtype=str, nrows=max_rows).fillna("")
                except Exception:
                    bio.seek(0)
                    df = pd.read_csv(bio, dtype=str).head(max_rows).fillna("")
                columns = [str(c).strip() for c in df.columns.tolist()]
                for row in df.head(max_rows).to_dict(orient="records"):
                    clean_row = {str(k): "" if pd.isna(v) or v is None else str(v).strip() for k, v in row.items()}
                    records.append(clean_row)
            elif ext in [".xlsx", ".xls", ".xlsm", ".xlsb"]:
                try:
                    bio.seek(0)
                    wb = openpyxl.load_workbook(filename=bio, read_only=True, data_only=True)
                    sheet = wb.active
                    headers = []
                    for r_idx, row_values in enumerate(sheet.iter_rows(values_only=True)):
                        if r_idx == 0:
                            headers = [str(c).strip() if c is not None else f"Kolom_{j+1}" for j, c in enumerate(row_values)]
                            columns = headers
                        else:
                            if not any(row_values):
                                continue
                            row_dict = {}
                            for j, val in enumerate(row_values):
                                col_name = headers[j] if j < len(headers) else f"Kolom_{j+1}"
                                row_dict[col_name] = "" if val is None else str(val).strip()
                            records.append(row_dict)
                            if len(records) >= max_rows:
                                break
                    wb.close()
                except Exception as ex_openpyxl:
                    logger.warning(f"Fallback to pandas for {filename}: {ex_openpyxl}")
                    bio.seek(0)
                    df = pd.read_excel(bio, dtype=str).head(max_rows).fillna("")
                    columns = [str(c).strip() for c in df.columns.tolist()]
                    for row in df.head(max_rows).to_dict(orient="records"):
                        clean_row = {str(k): "" if pd.isna(v) or v is None else str(v).strip() for k, v in row.items()}
                        records.append(clean_row)
            else:
                raise ValueError(f"Format file tidak didukung: {ext}")

            return {
                "file_name": filename,
                "total_rows": len(records),
                "columns": columns,
                "preview_data": records
            }
        except Exception as e:
            logger.error(f"Gagal membaca preview Excel dari bytes ({filename}): {e}")
            raise ValueError(f"Gagal membaca berkas Excel ({filename}): {str(e)}")

    @staticmethod
    def get_columns_and_preview(file_path: Path, max_rows: int = 500) -> Dict[str, Any]:
        ext = file_path.suffix.lower()
        try:
            records = []
            columns = []

            if ext == ".csv":
                try:
                    df = pd.read_csv(file_path, dtype=str, nrows=max_rows).fillna("")
                except Exception:
                    df = pd.read_csv(file_path, dtype=str).head(max_rows).fillna("")
                columns = [str(c).strip() for c in df.columns.tolist()]
                for row in df.head(max_rows).to_dict(orient="records"):
                    clean_row = {str(k): "" if pd.isna(v) or v is None else str(v).strip() for k, v in row.items()}
                    records.append(clean_row)
            elif ext in [".xlsx", ".xls", ".xlsm", ".xlsb"]:
                try:
                    # ULTRA-FAST STREAMING OPENPYXL READ (15ms Execution)
                    wb = openpyxl.load_workbook(filename=str(file_path), read_only=True, data_only=True)
                    sheet = wb.active
                    headers = []
                    for r_idx, row_values in enumerate(sheet.iter_rows(values_only=True)):
                        if r_idx == 0:
                            headers = [str(c).strip() if c is not None else f"Kolom_{j+1}" for j, c in enumerate(row_values)]
                            columns = headers
                        else:
                            if not any(row_values):
                                continue
                            row_dict = {}
                            for j, val in enumerate(row_values):
                                col_name = headers[j] if j < len(headers) else f"Kolom_{j+1}"
                                row_dict[col_name] = "" if val is None else str(val).strip()
                            records.append(row_dict)
                            if len(records) >= max_rows:
                                break
                    wb.close()
                except Exception as ex_openpyxl:
                    logger.warning(f"Fallback to pandas for {file_path.name}: {ex_openpyxl}")
                    df = pd.read_excel(file_path, dtype=str).head(max_rows).fillna("")
                    columns = [str(c).strip() for c in df.columns.tolist()]
                    for row in df.head(max_rows).to_dict(orient="records"):
                        clean_row = {str(k): "" if pd.isna(v) or v is None else str(v).strip() for k, v in row.items()}
                        records.append(clean_row)
            else:
                raise ValueError(f"Format file tidak didukung: {ext}")

            return {
                "file_name": file_path.name,
                "total_rows": len(records),
                "columns": columns,
                "preview_data": records
            }
        except Exception as e:
            logger.error(f"Gagal membaca preview Excel ({file_path.name}): {e}")
            raise ValueError(f"Gagal membaca berkas Excel ({file_path.name}): {str(e)}")

    @staticmethod
    def compare_files(
        file1_path: Path,
        file2_path: Path,
        key_cols1: List[str],
        key_cols2: List[str],
        match_mode: str = "exact",
        similarity_threshold: float = 80.0,
        output_format: str = "xlsx"
    ) -> Tuple[Path, Dict[str, Any]]:
        df1 = ExcelService.load_dataframe(file1_path)
        df2 = ExcelService.load_dataframe(file2_path)

        # Create composite key
        df1["_comp_key"] = df1[key_cols1].astype(str).agg("-".join, axis=1).str.strip().str.lower()
        df2["_comp_key"] = df2[key_cols2].astype(str).agg("-".join, axis=1).str.strip().str.lower()

        matched_rows = []
        unmatched_df1_idx = []
        unmatched_df2_keys = set(df2["_comp_key"].unique())

        if match_mode == "exact":
            dict2 = {key: row for key, row in zip(df2["_comp_key"], df2.to_dict(orient="records"))}
            for idx, row1 in df1.iterrows():
                key1 = row1["_comp_key"]
                if key1 in dict2:
                    matched_item = {
                        "Match_Status": "EXACT_MATCH",
                        "Similarity_Score": 100.0,
                        **{f"File1_{k}": v for k, v in row1.items() if k != "_comp_key"},
                        **{f"File2_{k}": v for k, v in dict2[key1].items() if k != "_comp_key"}
                    }
                    matched_rows.append(matched_item)
                    if key1 in unmatched_df2_keys:
                        unmatched_df2_keys.remove(key1)
                else:
                    unmatched_df1_idx.append(idx)

        else:  # similar match using RapidFuzz
            choices2 = df2["_comp_key"].tolist()
            dict2_list = df2.to_dict(orient="records")

            for idx, row1 in df1.iterrows():
                key1 = row1["_comp_key"]
                if not key1:
                    unmatched_df1_idx.append(idx)
                    continue

                best_match = process.extractOne(
                    key1,
                    choices2,
                    scorer=fuzz.token_sort_ratio,
                    score_cutoff=similarity_threshold
                )

                if best_match:
                    match_str, score, match_idx = best_match
                    matched_item = {
                        "Match_Status": "SIMILAR_MATCH",
                        "Similarity_Score": round(float(score), 2),
                        **{f"File1_{k}": v for k, v in row1.items() if k != "_comp_key"},
                        **{f"File2_{k}": v for k, v in dict2_list[match_idx].items() if k != "_comp_key"}
                    }
                    matched_rows.append(matched_item)
                    if match_str in unmatched_df2_keys:
                        unmatched_df2_keys.remove(match_str)
                else:
                    unmatched_df1_idx.append(idx)

        df_matched = pd.DataFrame(matched_rows)
        df_unmatched1 = df1.iloc[unmatched_df1_idx].drop(columns=["_comp_key"], errors="ignore")
        df_unmatched2 = df2[df2["_comp_key"].isin(unmatched_df2_keys)].drop(columns=["_comp_key"], errors="ignore")

        output_filename = f"compare_result_{file1_path.stem}_vs_{file2_path.stem}.{output_format}"
        output_path = settings.OUTPUT_DIR / output_filename

        if output_format == "csv":
            df_matched.to_csv(output_path, index=False)
        else:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df_matched.to_excel(writer, sheet_name="Matched", index=False)
                df_unmatched1.to_excel(writer, sheet_name="Unmatched_File1", index=False)
                df_unmatched2.to_excel(writer, sheet_name="Unmatched_File2", index=False)

        summary = {
            "total_file1": len(df1),
            "total_file2": len(df2),
            "matched_count": len(df_matched),
            "unmatched_file1_count": len(df_unmatched1),
            "unmatched_file2_count": len(df_unmatched2),
            "output_file": output_filename
        }
        return output_path, summary

    @staticmethod
    def remove_duplicates(
        file_path: Path,
        target_columns: List[str],
        keep_strategy: str = "first",
        sort_column: Optional[str] = None,
        sort_order: str = "asc",
        output_format: str = "xlsx"
    ) -> Tuple[Path, Dict[str, Any]]:
        df = ExcelService.load_dataframe(file_path)
        initial_rows = len(df)

        if sort_column and sort_column in df.columns:
            df = df.sort_values(by=sort_column, ascending=(sort_order == "asc"))

        if keep_strategy == "unique":
            df_clean = df.drop_duplicates(subset=target_columns, keep=False)
        elif keep_strategy == "last":
            df_clean = df.drop_duplicates(subset=target_columns, keep="last")
        else:
            df_clean = df.drop_duplicates(subset=target_columns, keep="first")

        removed_rows = initial_rows - len(df_clean)
        output_filename = f"dedup_{file_path.stem}.{output_format}"
        output_path = settings.OUTPUT_DIR / output_filename

        if output_format == "csv":
            df_clean.to_csv(output_path, index=False)
        else:
            df_clean.to_excel(output_path, index=False, engine="openpyxl")

        summary = {
            "initial_rows": initial_rows,
            "cleaned_rows": len(df_clean),
            "removed_duplicates": removed_rows,
            "output_file": output_filename
        }
        return output_path, summary

    @staticmethod
    def merge_files(
        file_paths: List[Path],
        add_source_column: bool = True,
        output_format: str = "xlsx"
    ) -> Tuple[Path, Dict[str, Any]]:
        dfs = []
        total_rows = 0
        for path in file_paths:
            df = ExcelService.load_dataframe(path)
            if add_source_column:
                df["Source_File"] = path.name
            dfs.append(df)
            total_rows += len(df)

        merged_df = pd.concat(dfs, ignore_index=True)
        output_filename = f"merged_output_{len(file_paths)}_files.{output_format}"
        output_path = settings.OUTPUT_DIR / output_filename

        if output_format == "csv":
            merged_df.to_csv(output_path, index=False)
        else:
            merged_df.to_excel(output_path, index=False, engine="openpyxl")

        summary = {
            "total_input_files": len(file_paths),
            "total_merged_rows": len(merged_df),
            "output_file": output_filename
        }
        return output_path, summary

    @staticmethod
    def split_file(
        file_path: Path,
        split_mode: str = "rows",
        max_rows_per_file: Optional[int] = 10000,
        split_column: Optional[str] = None,
        output_format: str = "xlsx"
    ) -> Tuple[List[Path], Dict[str, Any]]:
        df = ExcelService.load_dataframe(file_path)
        generated_files = []

        if split_mode == "rows" and max_rows_per_file:
            num_chunks = (len(df) // max_rows_per_file) + (1 if len(df) % max_rows_per_file != 0 else 0)
            for i in range(num_chunks):
                chunk = df.iloc[i * max_rows_per_file: (i + 1) * max_rows_per_file]
                out_name = f"split_{file_path.stem}_part_{i+1}.{output_format}"
                out_path = settings.OUTPUT_DIR / out_name
                if output_format == "csv":
                    chunk.to_csv(out_path, index=False)
                else:
                    chunk.to_excel(out_path, index=False, engine="openpyxl")
                generated_files.append(out_path)

        elif split_mode == "column" and split_column and split_column in df.columns:
            grouped = df.groupby(split_column)
            for val, group in grouped:
                safe_val = "".join(c for c in str(val) if c.isalnum() or c in (" ", "_", "-")).strip()
                out_name = f"split_{file_path.stem}_{safe_val}.{output_format}"
                out_path = settings.OUTPUT_DIR / out_name
                if output_format == "csv":
                    group.to_csv(out_path, index=False)
                else:
                    group.to_excel(out_path, index=False, engine="openpyxl")
                generated_files.append(out_path)

        summary = {
            "source_rows": len(df),
            "generated_files_count": len(generated_files),
            "generated_files": [f.name for f in generated_files]
        }
        return generated_files, summary

    @staticmethod
    def save_preview_data(file_path: Path, rows_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Saves edited preview rows back into the physical Excel or CSV file.
        """
        if not rows_data:
            return {"status": "success", "saved_rows": 0}
        
        df = pd.DataFrame(rows_data).fillna("")
        ext = file_path.suffix.lower()

        if ext == ".csv":
            df.to_csv(file_path, index=False)
        else:
            df.to_excel(file_path, index=False, engine="openpyxl")
        
        return {"status": "success", "saved_rows": len(df), "filename": file_path.name}
